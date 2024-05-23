"""Module for abstract serializer/unserializer base classes."""
import csv
import io
import json
from datetime import datetime
from random import choice
from string import ascii_lowercase, digits

import openpyxl
from bs4 import BeautifulSoup
from dateutil import parser
from dateutil.tz import gettz
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.contenttypes.models import ContentType
from django.contrib.gis.geos import Point
from django.core.files.storage import default_storage
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Count, F, Q, TextField
from django.db.models.functions import Cast
from django.http import Http404, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import redirect, render
from django.template.defaultfilters import slugify
from django.template.loader import render_to_string
from django.views import View
from django.views.generic import RedirectView
from reversion.views import RevisionMixin

from api.serializers import (
    CollegeListSerializer,
    FileSerializer,
    HomePageSerializer,
    PageSerializer,
    UniversityListSerializer,
)
from core.forms import (
    CareerForm,
    CollegeForm,
    LoginForm,
    MenuForm,
    PageForm,
    TileForm,
    UniversityForm,
)
from core.models import (
    Admin,
    AuditVersion,
    Career,
    College,
    CustomPage,
    File,
    Major,
    Menu,
    School,
    Tile,
    UniversityProfile,
    UserType,
)

# from reversion_compare.views import HistoryCompareDetailView


RANGE_LIST = [10, 25, 50, 100]


class BaseView(RevisionMixin, LoginRequiredMixin, View):
    """Login Decorator."""

    login_url = "/login/"
    redirect_field_name = "next"
    university = None

    university_check_failure_path = "404"

    def user_check_failed(self, request, *args, **kwargs):
        """sss."""
        return redirect(self.university_check_failure_path)

    def dispatch(self, request, *args, **kwargs):
        """Check user credentials."""
        if request.user.is_authenticated:
            university = request.session.get("university", None)
            admin = request.user.admin
            if not university:
                super_admin = request.user.is_staff or admin.role == "super_admin"
                if not super_admin and admin.universities.all().count() == 0:
                    return self.user_check_failed(request, *args, **kwargs)
            elif not admin.active:
                raise Http404
            elif (
                university
                and not UniversityProfile.objects.filter(uid=university).exists()
            ):
                return self.user_check_failed(request, *args, **kwargs)
            else:
                self.university = UniversityProfile.objects.get(uid=university)
            return super(BaseView, self).dispatch(request, *args, **kwargs)
        return redirect(self.login_url + "?next=" + request.get_full_path())


class SuperAdminBaseView(RevisionMixin, LoginRequiredMixin, View):
    """Login Decorator for super admin."""

    login_url = "/login/"
    redirect_field_name = "next"
    university = None

    university_check_failure_path = "404"

    def user_check_failed(self, request, *args, **kwargs):
        """sss."""
        return redirect(self.university_check_failure_path)

    def dispatch(self, request, *args, **kwargs):
        """Check user credentials."""
        if request.user.is_authenticated:
            admin = request.user.admin
            super_admin = request.user.is_staff or admin.role == "super_admin"
            if not super_admin:
                return self.user_check_failed(request, *args, **kwargs)
            return super(SuperAdminBaseView, self).dispatch(request, *args, **kwargs)
        return redirect(self.login_url + "?next=" + request.get_full_path())


class LogoutView(RedirectView):
    """Provides users the ability to logout."""

    url = "/login/"

    def get(self, request, *args, **kwargs):
        """Logout view."""
        logout(request)
        return super(LogoutView, self).get(request, *args, **kwargs)


class AdminDashBoard(BaseView):
    """Admin Dashboard View."""

    template_name = "superadmin/dashboard.html"

    def get(self, request, *args, **kwargs):
        """Render template view."""
        if request.user.is_authenticated:
            request.session["front_end"] = False
            admin = request.user.admin
            super_admin = request.user.is_staff or admin.role == "super_admin"
            if super_admin:
                universities = UniversityProfile.objects.filter(active=True)
            else:
                universities = admin.universities.filter(active=True)

            context = {"super_admin": super_admin, "universities": universities}
            return render(request, self.template_name, context)


class DashBoard(BaseView):
    """Dashboard View."""

    template_name = "login.html"

    def get(self, request, *args, **kwargs):
        """Render template view."""
        if request.user.is_authenticated:
            admin = request.user.admin
            university = self.university
            request.session["front_end"] = True

            if not admin.active:
                return HttpResponseRedirect("/login/")
            university_id = request.GET.get("university", None)
            if (
                university_id
                and UniversityProfile.objects.filter(uid=university_id).exists()
            ):
                university = UniversityProfile.objects.get(uid=university_id)
            if university and university.active:
                request.session["university"] = str(university.uid)
                request.session["university_logo"] = (
                    university.logo.url if university.logo else None
                )
                self.university = university
                return render(request, "home.html")
            if admin.role == "admin" and admin.universities.all().count() == 1:
                university = admin.universities.all().first()
                request.session["university"] = str(university.uid)
                request.session["university_logo"] = (
                    university.logo.url if university.logo else None
                )
                self.university = university
                return render(request, "home.html")
            if (
                request.user.is_staff
                or admin.role in ["super_admin", "admin"]
                or admin.universities.count() > 1
            ):
                return HttpResponseRedirect("/admin/")

        return render(request, self.template_name)


class LoginView(View):
    """View for login."""

    form_class = LoginForm
    template_name = "login.html"

    def get(self, request, *args, **kwargs):
        """Render Login form."""
        next_page = request.GET.get("next", None)
        if request.user and request.user.is_authenticated:
            admin = request.user.admin
            if not admin.active:
                return redirect("/404/")
            if next_page:
                return HttpResponseRedirect(next_page)
            return HttpResponseRedirect("/dashboard/")
        form = self.form_class()
        return render(request, self.template_name, {"form": form, "next": next_page})

    def post(self, request, *args, **kwargs):
        """Make user login."""
        next_page = request.GET.get("next", None)
        form = self.form_class(request.POST)
        if form.is_valid():
            email = form.data.get("email")
            password = form.data.get("password")
            username = User.objects.get(email__iexact=email).username
            user = authenticate(request, username=username, password=password)
            login(request, user)
            admin = user.admin
            if not admin.active:
                return redirect("/404/")
            if next_page:
                return HttpResponseRedirect(next_page)
            return HttpResponseRedirect("/dashboard/")

        return render(request, self.template_name, {"form": form, "next": next_page})


class ManageSchoolView(BaseView):
    """View for managing School."""

    template_name = "manage_school.html"

    def get(self, request, *args, **kwargs):
        """Render Template."""
        schools = School.objects.filter(university=self.university).order_by("name")
        if request.is_ajax():
            page = request.GET.get("page", 1)
            per_page = request.GET.get("per_page", 10)
            query = request.GET.get("query", None)
            if query:
                schools = schools.filter(name__icontains=query)

            paginator = Paginator(schools, per_page)

            try:
                schools = paginator.page(page)
            except (PageNotAnInteger, EmptyPage):
                if isinstance(e, PageNotAnInteger):
                    schools = paginator.page(1)
            else:
                schools = paginator.page(paginator.num_pages)
            # except PageNotAnInteger:
            #     schools = paginator.page(1)
            # except EmptyPage:
            #     schools = paginator.page(paginator.num_pages)
            # html = render_to_string('schools_list.html',
            #                         {'schools': schools,
            #                          'range_list': RANGE_LIST,
            #                          'query': query})
            # return HttpResponse(html)
        return render(request, self.template_name, {"schools": schools})

    def post(self, request, *args, **kwargs):
        """Save and Update School."""
        data = request.POST.copy()

        if "bulkDelete" in data:
            multiple_uids = data.getlist("multiple_uids[]", [])
            for school in School.objects.filter(uid__in=multiple_uids):
                school.delete()
            return HttpResponse("Deleted")
        if "csvfile" in request.FILES:
            xls_file = request.FILES["csvfile"]
            wb_obj = openpyxl.load_workbook(xls_file)
            ws = wb_obj.active
            duplicates = new_school = skipped = 0
            base_headers = [
                "school-name",
                "lea-name",
                "address-line-1",
                "address-line-2",
                "city",
                "state",
                "zip-code",
                "latitude",
                "longitude",
            ]
            header = [
                slugify(i.value)
                for i in tuple(ws.rows)[0]
                if slugify(i.value) in base_headers
            ]
            for row in ws.iter_rows(values_only=True):
                data = dict(zip(header, row))
                if "school-name" in data:
                    if slugify(data["school-name"]) == "school-name":
                        continue
                    if not data.get("school-name"):
                        skipped += 1
                        continue
                    created = error = False
                    try:
                        school, created = School.objects.get_or_create(
                            name=data.get("school-name").strip()
                            if data.get("school-name")
                            else None,
                            lea_name=data.get("lea-name").strip()
                            if data.get("lea-name")
                            else None,
                            address_line1=data.get("address-line-1").strip()
                            if data.get("address-line-1")
                            else "",
                            address_line2=data.get("address-line-2").strip()
                            if data.get("address-line-2")
                            else "",
                            city=data.get("city").strip() if data.get("city") else "",
                            state=data.get("state").strip()
                            if data.get("state")
                            else "",
                            zip_code=str(data.get("zip-code")).strip()
                            if data.get("zip-code")
                            else "",
                            university=self.university,
                            point=Point(
                                float(data.get("longitude", "").strip()),
                                float(data.get("latitude", "").strip()),
                                srid=4326,
                            )
                            if "latitude" and "longitude" in data
                            else None,
                        )
                    except Exception as e:
                        print(e)
                        error = True
                        skipped += 1
                        pass
                    if not error:
                        if created:
                            new_school += 1
                        else:
                            duplicates += 1

            return HttpResponse(
                json.dumps(
                    {
                        "result": "Success",
                        "duplicates": duplicates,
                        "created": new_school,
                        "errors": skipped,
                    }
                )
            )
        else:
            success = "true"
            message = "School Added Successfully"
            uid = data.get("uid", None)
            name = data.get("name", "").strip()
            lea_name = data.get("lea_name", "").strip()
            address_line1 = data.get("address_line1", "").strip()
            address_line2 = data.get("address_line2", "").strip()
            city = data.get("city", "").strip()
            state = data.get("state", "").strip()
            zip_code = data.get("zip_code", "").strip()
            if (
                not uid
                and School.objects.filter(
                    name=name,
                    lea_name=lea_name,
                    address_line1=address_line1,
                    address_line2=address_line2,
                    city=city,
                    state=state,
                    university=self.university,
                    zip_code=zip_code,
                ).exists()
            ):
                success = "false"
                message = "School with same details already exists"
                return HttpResponse(
                    json.dumps({"success": success, "message": message})
                )
            if uid and School.objects.filter(uid=uid).first():
                school = School.objects.get(uid=uid)
                school.name = name
                school.lea_name = lea_name
                school.address_line1 = address_line1
                school.address_line2 = address_line2
                school.city = city
                school.state = state
                school.zip_code = zip_code
                school.university = self.university
                school.save()
            else:
                school = School.objects.get_or_create(
                    name=name,
                    lea_name=lea_name,
                    address_line1=address_line1,
                    address_line2=address_line2,
                    city=city,
                    state=state,
                    university=self.university,
                    zip_code=zip_code,
                )[0]

        # lajdslkfjal

        return HttpResponse(json.dumps({"success": success, "message": message}))

    def delete(self, request, *args, **kwargs):
        """Deleted selected school."""
        uid = request.GET.get("uid", None)
        if uid and School.objects.filter(uid=uid).exists():
            School.objects.get(uid=uid).delete()
        return HttpResponse("Deleted")


class ManageMenuView(BaseView):
    """View for managing menu."""

    template_name = "manage_menu.html"

    def get(self, request, *args, **kwargs):
        """Render Template."""
        menus = Menu.objects.filter(university=self.university).order_by("order")
        home_page = CustomPage.objects.filter(
            home_page=True, university=self.university
        ).first()
        pages = (
            CustomPage.objects.filter(active=True, university=self.university)
            .exclude(uid=home_page.uid)
            .order_by("title")
        )
        old_order = dict(
            Menu.objects.filter(university=self.university)
            .annotate(str_id=Cast("uid", output_field=TextField()))
            .values_list("str_id", "order")
        )
        if request.is_ajax():
            html = render_to_string(
                "menu_list.html",
                {"menus": menus, "pages": pages, "old_order": json.dumps(old_order)},
            )
            return HttpResponse(html)
        return render(request, self.template_name, {"menus": menus, "pages": pages})

    def post(self, request, *args, **kwargs):
        """Save and Update Menu."""
        menu = Menu()
        data = request.POST.copy()
        uid = data.get("uid", None)
        if uid and Menu.objects.filter(university=self.university, uid=uid).first():
            menu = Menu.objects.filter(university=self.university, uid=uid).first()
            order = menu.order
        else:
            order = (
                Menu.objects.filter(university=self.university)
                .order_by("-order")
                .first()
                .order
                + 1
                if Menu.objects.filter(university=self.university)
                .order_by("-order")
                .exists()
                else 0
            )
        menu.order = order
        active = data.get("active")
        menu.active = True if active == "true" else False
        form = MenuForm(data, request.FILES, instance=menu)
        if data.get("type", None) == "order_update":
            order_data = json.loads(data.get("values"))
            for menu in Menu.objects.filter(uid__in=order_data.keys()):
                menu.order = order_data[str(menu.uid)]
                menu.university = self.university
                menu.save()
        elif data.get("type", None) == "menu_update":
            if form.is_valid():
                menu.university = self.university
                menu = form.save()
                if menu.url_type == "external":
                    menu.page = None
                    menu.key_name = None
                elif menu.url_type == "internal":
                    menu.url = None
                    menu.key_name = None
                elif menu.url_type == "pre_defined":
                    menu.page = None
                    menu.url = None
                menu.save()
            else:
                HttpResponse(form.errors)
        return HttpResponse("Success")

    def delete(self, request, *args, **kwargs):
        """Deleted selected menu."""
        uid = request.GET.get("uid", None)
        menu = Menu.objects.get(uid=uid)
        order = menu.order
        menu.delete()
        Menu.objects.filter(university=self.university, order__gte=order).update(
            order=F("order") - 1
        )
        return HttpResponse("Deleted")


class ManageHomePageView(BaseView):
    """View for managing Home Page."""

    template_name = "manage_home_page.html"

    def get(self, request, *args, **kwargs):
        """Render Template."""
        home_page = CustomPage.objects.filter(
            university=self.university, home_page=True
        ).first()
        print(home_page, self.university)
        pages = (
            CustomPage.objects.filter(university=self.university, active=True)
            .exclude(uid=home_page.uid)
            .order_by("title")
        )
        tiles = home_page.tiles.all()
        old_order = dict(
            tiles.annotate(str_id=Cast("uid", output_field=TextField())).values_list(
                "str_id", "order"
            )
        )
        context = {
            "home_page": home_page,
            "pages": pages,
            "tiles": tiles,
            "old_order": json.dumps(old_order),
            "video_id": home_page.video_id if home_page.video else None,
            "video_name": home_page.video.name if home_page.video else None,
            "video_url": home_page.video.cdn_url if home_page.video else None,
        }
        if request.is_ajax():
            if (
                request.GET.get("request_type")
                and request.GET.get("request_type") == "get_home_page"
            ):
                if request.GET.get("uid"):
                    home_page = CustomPage.objects.filter(
                        university=self.university, uid=request.GET.get("uid")
                    ).first()
                    serializer = HomePageSerializer(instance=home_page)
                return HttpResponse(json.dumps({"home_page": serializer.data}))
            html = render_to_string("tiles_list.html", context)
            return HttpResponse(html)
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Save and Update Home Page."""
        home_page = CustomPage.objects.filter(
            university=self.university, home_page=True
        ).first()
        tile = Tile()
        data = request.POST.copy()
        uid = data.get("uid", None)
        if uid and Tile.objects.filter(university=self.university, uid=uid).first():
            tile = Tile.objects.filter(university=self.university, uid=uid).first()
            order = tile.order
        else:
            order = (
                home_page.tiles.all().order_by("-order").first().order + 1
                if home_page.tiles.all().exists()
                else 0
            )
        tile.order = order
        active = data.get("active")
        tile.active = True if active == "true" else False
        form = TileForm(data, request.FILES, instance=tile)
        if data.get("type", None) == "order_update":
            order_data = json.loads(data.get("values"))
            for tile in Tile.objects.filter(uid__in=order_data.keys()):
                tile.order = order_data[str(tile.uid)]
                tile.save()
        elif data.get("type", None) == "tile_update":
            if form.is_valid():
                tile.university = self.university
                tile = form.save()
                if tile.url_type == "external":
                    tile.page = None
                if tile.url_type == "internal":
                    tile.url = None
                if not data.get("image_url", None) and not request.FILES.get("image"):
                    tile.image = None
                if not tile.active:
                    tile.modified_by_page = "deactive"
                else:
                    tile.modified_by_page = None
                tile.save()
                home_page.tiles.add(tile)
            else:
                HttpResponse(form.errors)
        elif data.get("type", None) == "home_page_update":
            video_title = data.get("banner_video_tile")
            home_page = CustomPage.objects.filter(
                university=self.university, home_page=True
            ).first()
            if home_page:
                video = home_page.video
                if video:
                    video.name = video_title
                    video.university = self.university
                    if request.FILES.get("thumbnail_image"):
                        video.thumbnail = request.FILES.get("thumbnail_image")
                    video.save()
        return HttpResponse("Success")

    def delete(self, request, *args, **kwargs):
        """Deleted selected Home Page."""
        uid = request.GET.get("uid", None)
        if Tile.objects.filter(uid=uid).exists():
            Tile.objects.get(uid=uid).delete()
        return HttpResponse("Deleted")


class FileView(BaseView):
    """View for managing File."""

    def get(self, request, *args, **kwargs):
        """Render Pages."""
        if request.is_ajax():
            if request.GET.get("uid"):
                video = File.objects.filter(
                    university=self.university, uid=request.GET.get("uid")
                ).first()
                if video:
                    serializer = FileSerializer(instance=video)
                    return HttpResponse(json.dumps({"video": serializer.data}))
        return HttpResponse(json.dumps({}))

    def post(self, request, *args, **kwargs):
        """Save a file."""
        file = request.FILES["image"]
        file_name = default_storage.save(file.name, file)
        file = default_storage.open(file_name)
        file_url = default_storage.url(file_name)
        full_path = "%(scheme)s://%(host)s%(file_url)s" % {
            "scheme": request.scheme,
            "host": request.get_host(),
            "file_url": file_url,
        }
        return HttpResponse(json.dumps({"data": full_path}))

    def delete(self, request, *args, **kwargs):
        """Deleted selected Home Page."""
        uid = request.GET.get("uid", None)
        if uid and File.objects.filter(uid=uid).exists():
            File.objects.get(uid=uid).delete()
        return HttpResponse("Deleted")


class ManageCustomView(BaseView):
    """View for managing Pages."""

    template_name = "manage_custom_page.html"

    def get(self, request, *args, **kwargs):
        """Render Pages."""
        pages = CustomPage.objects.filter(
            university=self.university, home_page=False
        ).order_by("title")
        if request.is_ajax():
            if request.is_ajax():
                if (
                    request.GET.get("request_type")
                    and request.GET.get("request_type") == "get_page"
                ):
                    if request.GET.get("uid"):
                        page = CustomPage.objects.get(
                            university=self.university, uid=request.GET.get("uid")
                        )
                        data = PageSerializer(instance=page, html_content=False).data
                        data["related_tiles"] = page.related_tiles
                        return HttpResponse(json.dumps({"page": data}))
            page = request.GET.get("page", 1)
            per_page = request.GET.get("per_page", 10)
            query = request.GET.get("query", None)
            if query:
                pages = pages.filter(title__icontains=query)

            paginator = Paginator(pages, per_page)

            try:
                pages = paginator.page(page)

            except (PageNotAnInteger, EmptyPage) as e:
                if isinstance(e, PageNotAnInteger):
                    pages = paginator.page(1)
                else:
                    pages = paginator.page(paginator.num_pages)

            html = render_to_string(
                "pages_list.html",
                {"pages": pages, "range_list": RANGE_LIST, "query": query},
            )
            return HttpResponse(html)
            # except PageNotAnInteger:
            #     pages = paginator.page(1)
            # except EmptyPage:
            #     pages = paginator.page(paginator.num_pages)
            # html = render_to_string('pages_list.html',
            #                         {'pages': pages,
            #                          'range_list': RANGE_LIST,
            #                          'query': query})
            # return HttpResponse(html)
        return render(request, self.template_name, {"pages": pages})

    def post(self, request, *args, **kwargs):
        """Save and Update Page."""
        page = CustomPage()
        data = request.POST.copy()

        if "bulkDelete" in data:
            multiple_uids = data.getlist("multiple_uids[]", [])
            for page in CustomPage.objects.filter(uid__in=multiple_uids):
                page.delete()
                Tile.objects.filter(page=page, university=page.university).delete()
            return HttpResponse("Deleted")
        uid = data.get("uid", None)
        if (
            uid
            and CustomPage.objects.filter(university=self.university, uid=uid).first()
        ):
            page = CustomPage.objects.filter(
                university=self.university, uid=uid
            ).first()

        active = data.get("active")
        page.active = True if active == "true" else False
        form = PageForm(data, instance=page)
        if form.is_valid():
            page.university = self.university
            page = form.save()
            if (
                data.get("file_id", None)
                and File.objects.filter(
                    university=self.university, uid=data.get("file_id")
                ).exists()
            ):
                page.video = File.objects.get(uid=data.get("file_id"))
            page.save()
            for tile in Tile.objects.filter(page=page, university=page.university):
                if page.active:
                    if tile.modified_by_page == "modified":
                        tile.active = True
                        tile.save()
                else:
                    tile.active = False
                    if not tile.modified_by_page:
                        tile.modified_by_page = "modified"
                tile.save()

        else:
            return HttpResponse(form.errors)

        video_title = data.get("banner_video_title")
        video = page.video
        if video:
            if video.name != video_title:
                video.name = video_title
            video.university = self.university
            if request.FILES.get("thumbnail_image"):
                video.thumbnail = request.FILES.get("thumbnail_image")
            video.save()
        return HttpResponse("Success")

    def delete(self, request, *args, **kwargs):
        """Deleted selected Page."""
        uid = request.GET.get("uid", None)
        if CustomPage.objects.filter(uid=uid).exists():
            page = CustomPage.objects.get(uid=uid)
            Tile.objects.filter(page=page, university=page.university).delete()
            page.delete()
        return HttpResponse("Deleted")


class ManageMajorsView(BaseView):
    """View for managing Pages."""

    template_name = "manage_majors.html"

    def get(self, request, *args, **kwargs):
        """Render Pages."""
        majors = Major.objects.filter(university=self.university).order_by("title")
        if request.is_ajax():
            page = request.GET.get("page", 1)
            per_page = request.GET.get("per_page", 10)
            query = request.GET.get("query", None)
            if query:
                majors = majors.filter(title__icontains=query)

            paginator = Paginator(majors, per_page)

            try:
                majors = paginator.page(page)
            except (PageNotAnInteger, EmptyPage) as e:
                if isinstance(e, PageNotAnInteger):
                    majors = paginator.page(1)
                else:
                    majors = paginator.page(paginator.num_pages)

            html = render_to_string(
                "majors_list.html",
                {"majors": majors, "range_list": RANGE_LIST, "query": query},
            )
            return HttpResponse(html)
            # except PageNotAnInteger:
            #     majors = paginator.page(1)
            # except EmptyPage:
            #     majors = paginator.page(paginator.num_pages)
            # html = render_to_string('majors_list.html',
            #                         {'majors': majors,
            #                          'range_list': RANGE_LIST,
            #                          'query': query})
            # return HttpResponse(html)
        colleges = College.objects.filter(university=self.university).order_by("name")
        serializer = CollegeListSerializer(colleges, many=True)
        return render(request, self.template_name, {"colleges": serializer.data})

    def post(self, request, *args, **kwargs):
        """Save and Update Major."""
        data = request.POST.copy()
        if "bulkDelete" in data:
            multiple_uids = data.getlist("multiple_uids[]", [])
            for major in Major.objects.filter(uid__in=multiple_uids):
                major.delete()
            return HttpResponse("Deleted")
        active = data.get("active")
        if (
            data.get("uid", None)
            and Major.objects.filter(
                university=self.university, uid=data["uid"]
            ).first()
        ):
            major = Major.objects.filter(
                university=self.university, uid=data["uid"]
            ).first()
        else:
            major = Major.objects.get_or_create(
                university=self.university,
                title=data.get("title"),
                description=data.get("description"),
            )[0]
        major.active = True if active == "true" else False
        major.save()
        colleges = []
        if (
            data.get("colleges", None)
            and College.objects.filter(
                university=self.university, uid__in=data["colleges"].split(",")
            ).exists()
        ):
            colleges = College.objects.filter(
                university=self.university, uid__in=data["colleges"].split(",")
            )
        major.college_set.set(colleges)
        return HttpResponse("Success")

    def delete(self, request, *args, **kwargs):
        """Deleted selected Page."""
        uid = request.GET.get("uid", None)
        major = Major.objects.filter(uid=uid)
        major.delete()
        return HttpResponse("Deleted")


class ManageCollegesView(BaseView):
    """View for managing Colleges."""

    template_name = "manage_colleges.html"

    def get(self, request, *args, **kwargs):
        """Render Pages."""
        College.objects.filter(university=self.university, active=False).delete()
        queryset = College.objects.filter(university=self.university).order_by("name")
        page = request.GET.get("page", 1)
        per_page = request.GET.get("per_page", 10)
        query = request.GET.get("query", "")
        if query:
            queryset = queryset.filter(name__icontains=query)

        paginator = Paginator(queryset, per_page)

        try:
            queryset = paginator.page(page)
        except (PageNotAnInteger, EmptyPage) as e:
            if isinstance(e, PageNotAnInteger):
                queryset = paginator.page(1)
        else:
            queryset = paginator.page(paginator.num_pages)

        return render(
            request,
            self.template_name,
            {
                "colleges": queryset,
                "range_list": RANGE_LIST,
                "query": query,
                "per_page": per_page,
            },
        )
        # except PageNotAnInteger:
        #     queryset = paginator.page(1)
        # except EmptyPage:
        #     queryset = paginator.page(paginator.num_pages)
        # return render(request, self.template_name,
        #               {'colleges': queryset,
        #                'range_list': RANGE_LIST,
        #                'query': query,
        #                'per_page': per_page})

    def post(self, request, *args, **kwargs):
        data = request.POST.copy()
        if "bulkDelete" in data:
            multiple_uids = data.getlist("multiple_uids[]", [])
            for college in College.objects.filter(uid__in=multiple_uids):
                college.delete()
        return HttpResponse("Deleted")

    def delete(self, request, *args, **kwargs):
        """Deleted selected College."""
        uid = request.GET.get("uid", None)
        if College.objects.filter(uid=uid).exists():
            College.objects.get(uid=uid).delete()
        return HttpResponse("Deleted")


class CollegeDetailView(BaseView):
    """Manage a single College."""

    template_name = "college_detail.html"

    def get(self, request, college_uid=None):
        """Get the details of home page."""
        form = CollegeForm()
        college = None
        if not college_uid:
            college = College(active=False)
            college.university = self.university
            college.save()
            return redirect("college-detail", str(college.uid))

        tiles = None
        old_order = {}
        home_page = CustomPage.objects.filter(
            university=self.university, home_page=True
        ).first()
        pages = (
            CustomPage.objects.filter(university=self.university, active=True)
            .exclude(uid=home_page.uid)
            .order_by("title")
        )
        majors = Major.objects.filter(active=True, university=self.university)
        if (
            college_uid
            and College.objects.filter(
                university=self.university, uid=college_uid
            ).exists()
        ):
            college = College.objects.filter(
                university=self.university, uid=college_uid
            ).first()
            form = CollegeForm(instance=college)
            tiles = college.tiles.all().order_by("order")
            old_order = dict(
                tiles.annotate(
                    str_id=Cast("uid", output_field=TextField())
                ).values_list("str_id", "order")
            )
        context = {
            "majors": majors,
            "pages": pages,
            "tiles": tiles,
            "form": form,
            "college": college,
            "old_order": json.dumps(old_order),
            "college_uid": college_uid,
            "video_id": college.video_id if college and college.video else None,
            "video_name": college.video.name if college and college.video else None,
            "video_url": college.video.cdn_url if college and college.video else None,
        }

        if request.is_ajax():
            html = render_to_string("college_tiles_list.html", context)
            return HttpResponse(html)
        return render(request, self.template_name, context)

    def post(self, request, college_uid):
        """Create a new College."""
        data = request.POST.copy()
        college = College.objects.filter(
            university=self.university, uid=college_uid
        ).first()

        if college_uid:
            college = College.objects.filter(
                university=self.university, uid=college_uid
            ).first()
            if "csvfile" in request.FILES:
                xls_file = request.FILES["csvfile"]
                wb_obj = openpyxl.load_workbook(xls_file)
                ws = wb_obj.active
                new_majors = duplicates = 0
                for row in ws.iter_rows(values_only=True):
                    if row[0]:
                        major_title = row[0].strip()
                        major, created = Major.objects.get_or_create(
                            title=major_title, university=self.university
                        )
                        major.active = True
                        major.save()
                        college.majors.add(major)
                        if created:
                            new_majors += 1
                        else:
                            duplicates += 1
                return HttpResponse(
                    json.dumps(
                        {
                            "result": "Success",
                            "duplicates": duplicates,
                            "created": new_majors,
                        }
                    )
                )
            if data.get("type") == "add_major":
                major_title = data.get("major_title", None)
                if major_title:
                    major = Major.objects.get_or_create(
                        title=major_title, university=self.university
                    )[0]
                    major.active = True
                    major.save()
                    college.majors.add(major)
                return HttpResponse(json.dumps('{result: "Success"}'))
            elif data.get("type", None) == "tile_update":
                tile = Tile()
                tile_uid = data.get("tile_uid", None)
                if (
                    tile_uid
                    and Tile.objects.filter(
                        university=self.university, uid=tile_uid
                    ).first()
                ):
                    tile = Tile.objects.filter(
                        university=self.university, uid=tile_uid
                    ).first()
                    order = tile.order
                else:
                    order = (
                        college.tiles.all().all().order_by("-order").first().order + 1
                        if college.tiles.all().exists()
                        else 1
                    )
                tile.order = order
                active = data.get("tile_active")
                tile.active = True if active == "true" else False
                tile_form = TileForm(data, request.FILES, instance=tile)
                if tile_form.is_valid():
                    tile.university = self.university
                    tile = tile_form.save()
                    if tile.url_type == "external":
                        tile.page = None
                    if tile.url_type == "internal":
                        tile.url = None
                    if not data.get("image_url", None) and not request.FILES.get(
                        "image"
                    ):
                        tile.image = None
                    if not tile.active:
                        tile.modified_by_page = "deactive"
                    else:
                        tile.modified_by_page = None
                    tile.save()
                    college.tiles.add(tile)
                else:
                    return HttpResponse(json.dumps('{result: "Error"}'))
                return HttpResponse(json.dumps('{result: "Success"}'))
            elif data.get("type", None) == "order_update":
                order_data = json.loads(data.get("values"))
                for tile in Tile.objects.filter(uid__in=order_data.keys()):
                    tile.order = order_data[str(tile.uid)]
                    tile.university = self.university
                    tile.save()
        if "latitude" in data and "longitude" in data:
            data["location"] = Point(
                float(data.get("longitude")), float(data.get("latitude")), srid=4326
            )
        college.university = self.university
        data["university"] = str(self.university.uid)
        form = CollegeForm(data, request.FILES, instance=college)
        if form.is_valid():
            college = form.save()
            college.active = True
            college.save()
        else:
            return JsonResponse(form.errors.as_json(), safe=False)
        if (
            data.get("file_id", None)
            and File.objects.filter(
                university=self.university, uid=data.get("file_id")
            ).exists()
        ):
            video = File.objects.get(
                university=self.university, uid=data.get("file_id")
            )
            if data.get("video_title"):
                video.name = data.get("video_title")
            if request.FILES.get("thumbnail_image"):
                video.thumbnail = request.FILES.get("thumbnail_image")
            video.save()
            college.video = video
            college.save()

        majors = []
        if (
            data.get("major_ids", None)
            and Major.objects.filter(uid__in=data["major_ids"].split(",")).exists()
        ):
            majors = Major.objects.filter(
                university=self.university, uid__in=data["major_ids"].split(",")
            )
        college.majors.set(majors)
        majors = Major.objects.filter(university=self.university, active=True)
        home_page = CustomPage.objects.filter(
            university=self.university, home_page=True
        ).first()
        pages = (
            CustomPage.objects.filter(university=self.university, active=True)
            .exclude(uid=home_page.uid)
            .order_by("title")
        )
        tiles = college.tiles.all()
        context = {
            "majors": majors,
            "pages": pages,
            "tiles": tiles,
            "form": form,
            "college": college,
            "college_uid": college_uid,
            "video_id": college.video_id if college.video else None,
            "video_name": college.video.name if college.video else None,
            "video_url": college.video.cdn_url if college.video else None,
        }
        return render(request, self.template_name, context)


class EditProfileView(BaseView):
    """View for managing profile menu."""

    template_name = "superadmin/profile.html"

    def get(self, request, *args, **kwargs):
        """Render Template."""
        base_template = (
            "base.html" if request.session["front_end"] else "superadmin/base.html"
        )
        return render(request, self.template_name, {"base_template": base_template})

    def post(self, request, *args, **kwargs):
        """Save and Update Menu."""
        data = request.POST.copy()
        user = request.user
        if data.get("old_password") and not user.check_password(
            data.get("old_password")
        ):
            return HttpResponse(
                json.dumps(
                    {"old_password": "failed", "message": "Incorrect Old Password"}
                )
            )
        elif data.get("old_password") and user.check_password(data.get("old_password")):
            user.set_password(data.get("new_password"))
            user.save()
            logout(request)
            return HttpResponse(
                json.dumps(
                    {
                        "success": "true",
                        "message": "\
            Password updated successfully. Please login to continue again.",
                    }
                )
            )
        if data.get("first_name"):
            user.first_name = data.get("first_name")
        if data.get("last_name"):
            user.last_name = data.get("last_name")
        user.save()

        return HttpResponse(
            json.dumps(
                {"success": "true", "message": "Profile details updated successfully."}
            )
        )


class ManageUniversityView(SuperAdminBaseView):
    """View for managing university."""

    template_name = "superadmin/manage_universities.html"

    def get(self, request, *args, **kwargs):
        """Render Template."""
        queryset = UniversityProfile.objects.all().order_by("name")
        page = request.GET.get("page", 1)
        per_page = request.GET.get("per_page", 10)

        query = request.GET.get("query", "")
        if query:
            queryset = queryset.filter(name__icontains=query)

        paginator = Paginator(queryset, per_page)

        try:
            queryset = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            if isinstance(e, PageNotAnInteger):
                queryset = paginator.page(1)
        else:
            queryset = paginator.page(paginator.num_pages)

        if request.is_ajax():
            html = render_to_string(
                "superadmin/university_list.html",
                {"universities": queryset, "range_list": RANGE_LIST, "query": query},
            )
            return HttpResponse(html)
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        """Save and Update University."""
        data = request.POST.copy()
        uid = data.get("uid", None)
        if uid and UniversityProfile.objects.filter(uid=uid).first():
            university = UniversityProfile.objects.filter(uid=uid).first()
        else:
            university = UniversityProfile()
        active = data.get("active")
        university.active = True if active == "true" else False
        form = UniversityForm(data, request.FILES, instance=university)
        if form.is_valid():
            university = form.save()
            university.save()
            Admin.objects.annotate(num_universities=Count("universities")).filter(
                universities=university, num_universities=1
            ).update(active=university.active)
        else:
            HttpResponse(form.errors)
        if not CustomPage.objects.filter(
            university=university, home_page=True
        ).exists():
            home_page = CustomPage(
                university=university, home_page=True, active=True, title="Home Page"
            )
            home_page.save()
        for order, u_type in enumerate(["Student", "Teacher", "Parent", "Other"]):
            if not UserType.objects.filter(
                title=u_type, university=university
            ).exists():
                user_type = UserType(title=u_type, university=university, order=order)
                user_type.save()

        return HttpResponse(json.dumps({"success": "true", "message": "updated"}))

    def delete(self, request, *args, **kwargs):
        """Deleted selected menu."""
        uid = request.GET.get("uid", None)
        university = UniversityProfile.objects.get(uid=uid)
        university.active = False if university.active else True
        university.save()
        Admin.objects.annotate(num_universities=Count("universities")).filter(
            universities=university, num_universities=1
        ).update(active=university.active)
        return HttpResponse(json.dumps({"success": "true", "message": "deleted"}))


class ManageAPICredentialView(SuperAdminBaseView):
    """View for managing api credentials."""

    def generate_random_username(
        self, length=16, chars=ascii_lowercase + digits, split=4, delimiter="-"
    ):
        """Generate random username."""
        username = "".join([choice(chars) for i in range(length)])
        if split:
            username = delimiter.join(
                [
                    username[start : start + split]
                    for start in range(0, len(username), split)
                ]
            )
        try:
            User.objects.get(username=username)
            return self.generate_random_username(
                length=length, chars=chars, split=split, delimiter=delimiter
            )
        except User.DoesNotExist:
            return username

    def get(self, request, *args, **kwargs):
        """Render Template."""
        university_id = request.GET.get("university", None)
        university = UniversityProfile.objects.filter(uid=university_id).first()
        credentials = Admin.objects.filter(
            universities=university, role="silent_login"
        ).values("user__username", "user_id", "silent_login_password")
        html = render_to_string(
            "superadmin/credentials_list.html", {"credentials": credentials}
        )
        return HttpResponse(json.dumps({"success": "true", "html": html}))

    def post(self, request, *args, **kwargs):
        """Save and Update University."""
        data = request.POST.copy()
        uid = data.get("uid", None)
        if uid and UniversityProfile.objects.filter(uid=uid).first():
            university = UniversityProfile.objects.filter(uid=uid).first()
        else:
            university = UniversityProfile()
        active = data.get("active")
        university.active = True if active == "true" else False
        form = UniversityForm(data, request.FILES, instance=university)
        if form.is_valid():
            university = form.save()
            university.save()
        if (
            Admin.objects.filter(universities=university, role="silent_login").count()
            < 2
        ):
            username = self.generate_random_username(length=8)
            password = self.generate_random_username(length=16, delimiter="")
            api_user = User.objects.create(username=username)
            api_user.set_password(password)
            api_user.first_name = university.name
            api_user.last_name = datetime.utcnow().date()
            api_user.save()
            admin = Admin(
                user=api_user,
                role="silent_login",
                active=True,
                silent_login_password=password,
            )
            admin.save()
            admin.universities.add(university)

        credentials = Admin.objects.filter(
            universities=university, role="silent_login"
        ).values("user__username", "user_id", "silent_login_password")
        html = render_to_string(
            "superadmin/credentials_list.html", {"credentials": credentials}
        )
        return HttpResponse(
            json.dumps({"success": "true", "html": html, "uid": str(university.uid)})
        )

    def delete(self, request, *args, **kwargs):
        """Deleted selected menu."""
        uid = request.GET.get("university", None)
        university = UniversityProfile.objects.get(uid=uid)
        user_id = request.GET.get("user_id", None)
        if user_id and User.objects.filter(id=user_id).exists():
            user = User.objects.get(id=user_id)
            Admin.objects.get(user=user).delete()
            User.objects.get(id=user_id).delete()

        credentials = Admin.objects.filter(
            universities=university, role="silent_login"
        ).values("user__username", "user_id", "silent_login_password")
        html = render_to_string(
            "superadmin/credentials_list.html", {"credentials": credentials}
        )
        return HttpResponse(json.dumps({"success": "true", "html": html}))


class ManageAdminView(SuperAdminBaseView):
    """View for managing admin."""

    template_name = "superadmin/manage_admins.html"

    def send_verification_email(self, request, email):
        """Send verification email to admin."""
        reset_form = PasswordResetForm({"email": email})
        assert reset_form.is_valid()

        reset_form.save(
            request=request,
            use_https=True,
            from_email=settings.DEFAULT_FROM_EMAIL,
            email_template_name="registration/invitation_link.html",
        )
        return None

    def get(self, request, *args, **kwargs):
        """Render Template."""
        queryset = (
            Admin.objects.all().exclude(role="silent_login").order_by("user__username")
        )
        page = request.GET.get("page", 1)
        per_page = request.GET.get("per_page", 10)

        query = request.GET.get("query", "")
        if query:
            queryset = queryset.filter(
                Q(user__first_name__icontains=query)
                | Q(user__last_name__icontains=query)
                | Q(user__username__icontains=query)
                | Q(user__email__icontains=query)
            )

        paginator = Paginator(queryset, per_page)
        try:
            queryset = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            if isinstance(e, PageNotAnInteger):
                queryset = paginator.page(1)
            else:
                queryset = paginator.page(paginator.num_pages)
        # try:
        #     queryset = paginator.page(page)
        # except PageNotAnInteger:
        #     queryset = paginator.page(1)
        # except EmptyPage:
        #     queryset = paginator.page(paginator.num_pages)

        if request.is_ajax():
            html = render_to_string(
                "superadmin/admin_list.html",
                {"admins": queryset, "range_list": RANGE_LIST, "query": query},
            )
            return HttpResponse(html)
        universities = UniversityProfile.objects.filter(active=True)
        serializer = UniversityListSerializer(universities, many=True)
        return render(request, self.template_name, {"universities": serializer.data})

    def post(self, request, *args, **kwargs):
        """Save and Update University."""
        data = request.POST.copy()
        uid = data.get("uid", None)
        first_name = data.get("first_name")
        last_name = data.get("last_name")
        email = data.get("email")
        superadmin = data.get("superadmin", "off")
        active = data.get("active")
        university_ids = data.get("university_ids")
        resend_verification = data.get("resend_verification", "false")
        if resend_verification and resend_verification == "true":
            self.send_verification_email(request, email)
            return HttpResponse(
                json.dumps(
                    {"success": "true", "message": "Verification link resend to admin"}
                )
            )
        if uid and Admin.objects.filter(uid=uid).exists():
            admin = Admin.objects.get(uid=uid)
            if email and User.objects.filter(email=email).exists():
                user = User.objects.filter(email=email).first()
                if admin.user != user:
                    return HttpResponse(
                        json.dumps(
                            {
                                "success": "false",
                                "message": "User with this email already exists.",
                            }
                        )
                    )
        else:
            if email and User.objects.filter(email__iexact=email).exists():
                return HttpResponse(
                    json.dumps(
                        {
                            "success": "false",
                            "message": "User with this email already exists.",
                        }
                    )
                )
            else:
                user = User.objects.create_user(
                    username=email,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password="123456",
                    is_active=True,
                )
                user.save()
                admin = Admin(user=user)
                admin.save()
                self.send_verification_email(request, email)

        user = admin.user
        if first_name:
            user.first_name = first_name
        if last_name:
            user.last_name = last_name
        if email:
            user.email = email
        if active:
            admin.active = True if active == "true" else False
        admin.role = "super_admin" if superadmin == "on" else "admin"
        user.save()
        admin.save()

        universities = []
        if (
            data.get("university_ids", None)
            and UniversityProfile.objects.filter(
                uid__in=university_ids.split(",")
            ).exists()
        ):
            universities = UniversityProfile.objects.filter(
                uid__in=university_ids.split(",")
            )
        admin.universities.set(universities)
        return HttpResponse(json.dumps({"success": "true", "message": "updated"}))

    def delete(self, request, *args, **kwargs):
        """Deleted selected menu."""
        uid = request.GET.get("uid", None)
        Admin.objects.get(uid=uid).delete()
        return HttpResponse(json.dumps({"success": "true", "message": "deleted"}))


class ManageLogView(SuperAdminBaseView):
    """View for managing logs."""

    template_name = "superadmin/manage_logs.html"

    def get(self, request, *args, **kwargs):
        """Render Template."""
        queryset = (
            AuditVersion.objects.all()
            .order_by("-revision__date_created")
            .exclude(content_type=ContentType.objects.get_for_model(UserType))
        )
        page = request.GET.get("page", 1)
        per_page = request.GET.get("per_page", 10)
        start_date = request.GET.get("start_date", None)
        end_date = request.GET.get("end_date", None)

        filter_type = request.GET.get("filter_type", "All")

        query = request.GET.get("query", "")
        export_csv = request.GET.get("export_csv", False)
        selected_date_range = request.GET.get("selected_date_range", None)
        if query:
            university = UniversityProfile.objects.filter(name__icontains=query).first()
            users_list = queryset.filter(
                Q(revision__user__first_name__icontains=query)
                | Q(revision__user__last_name__icontains=query)
                | Q(revision__user__username__icontains=query)
                | Q(revision__user__email__icontains=query)
            )
            if university:
                universities_list = [
                    item.pk
                    for item in queryset
                    if item.object
                    and hasattr(item.object, "university")
                    and item.object.university == university
                ]
                queryset = users_list | queryset.filter(pk__in=universities_list)
            else:
                queryset = users_list
        if start_date and end_date:
            add_default_tz = lambda x, tzinfo: x.replace(tzinfo=x.tzinfo or tzinfo)
            start_date = add_default_tz(parser.parse(start_date), gettz("UTC"))
            end_date = add_default_tz(parser.parse(end_date), gettz("UTC"))
            if start_date != end_date:
                queryset = queryset.filter(revision__date_created__gte=start_date)
            else:
                queryset = queryset.filter(
                    revision__date_created__gte=start_date,
                    revision__date_created__lte=end_date,
                )

        if filter_type:
            if filter_type == "Deleted":
                queryset = [item for item in queryset if item.status == "deleted"]
            elif filter_type == "Created":
                queryset = [item for item in queryset if item.status == "created"]
            elif filter_type == "Modified":
                queryset = [item for item in queryset if item.status == "modified"]

        if export_csv:
            filename = "%s.csv" % datetime.now().strftime("%m/%d/%Y-%H:%M:%S")

            buffer = io.StringIO()
            wr = csv.writer(buffer, quoting=csv.QUOTE_ALL)
            for row in queryset:
                wr.writerow(
                    [
                        row.revision.date_created.strftime("%m/%d/%Y-%H:%M:%S"),
                        row.user.get_full_name(),
                        row.status,
                        row.content_type.name,
                        row.object_repr,
                        "-- ".join(
                            [
                                "%s: %s"
                                % (
                                    c["field"].name,
                                    BeautifulSoup(
                                        c["diff"], "html.parser"
                                    ).text.replace("\n", "  "),
                                )
                                for c in row.changes
                            ]
                        ),
                    ]
                )

            buffer.seek(0)
            response = HttpResponse(buffer, content_type="text/csv")
            response["Content-Disposition"] = "attachment; filename=%s" % filename

            return response
        paginator = Paginator(queryset, per_page)
        try:
            queryset = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            if isinstance(e, PageNotAnInteger):
                queryset = paginator.page(1)
            else:
                queryset = paginator.page(paginator.num_pages)
        # try:
        #     queryset = paginator.page(page)
        # except PageNotAnInteger:
        #     queryset = paginator.page(1)
        # except EmptyPage:
        #     queryset = paginator.page(paginator.num_pages)
        if request.is_ajax():
            html = render_to_string(
                "superadmin/logs_list.html",
                {
                    "logs": queryset,
                    "range_list": RANGE_LIST,
                    "filter_types": ["All", "Created", "Modified", "Deleted"],
                    "selected_filter": filter_type,
                    "query": query,
                    "selected_date_range": selected_date_range,
                },
            )
            return HttpResponse(html)
        return render(request, self.template_name)


class ManageCareersView(BaseView):
    """View for managing Colleges."""

    template_name = "manage_careers.html"

    def get(self, request, *args, **kwargs):
        """Render Pages."""
        Career.objects.filter(university=self.university, active=False).delete()
        queryset = Career.objects.filter(university=self.university).order_by("career")
        page = request.GET.get("page", 1)
        per_page = request.GET.get("per_page", 10)
        query = request.GET.get("query", "")
        if query:
            queryset = queryset.filter(career__icontains=query)

        paginator = Paginator(queryset, per_page)
        try:
            queryset = paginator.page(page)
        except (PageNotAnInteger, EmptyPage) as e:
            if isinstance(e, PageNotAnInteger):
                queryset = paginator.page(1)
            else:
                queryset = paginator.page(paginator.num_pages)

        return render(
            request,
            self.template_name,
            {
                "careers": queryset,
                "range_list": RANGE_LIST,
                "query": query,
                "per_page": per_page,
            },
        )
        # try:
        #     queryset = paginator.page(page)
        # except PageNotAnInteger:
        #     queryset = paginator.page(1)
        # except EmptyPage:
        #     queryset = paginator.page(paginator.num_pages)
        # return render(request, self.template_name,
        #               {'careers': queryset,
        #                'range_list': RANGE_LIST,
        #                'query': query,
        #                'per_page': per_page})

    def post(self, request, *args, **kwargs):
        data = request.POST.copy()
        if "bulkDelete" in data:
            multiple_uids = data.getlist("multiple_uids[]", [])
            for career in Career.objects.filter(uid__in=multiple_uids):
                career.delete()
        return HttpResponse("Deleted")

    def delete(self, request, *args, **kwargs):
        """Deleted selected College."""
        uid = request.GET.get("uid", None)
        if Career.objects.filter(uid=uid).exists():
            Career.objects.get(uid=uid).delete()
        return HttpResponse("Deleted")


class CareerDetailView(BaseView):
    """Manage a single College."""

    template_name = "career_detail.html"

    def get(self, request, career_uid=None):
        """Get the details of home page."""
        form = CareerForm()
        career = None
        if not career_uid:
            career = Career(active=False)
            career.university = self.university
            career.save()
            return redirect("career-detail", str(career.uid))

        tiles = None
        old_order = {}

        if (
            career_uid
            and Career.objects.filter(
                university=self.university, uid=career_uid
            ).exists()
        ):
            career = Career.objects.filter(
                university=self.university, uid=career_uid
            ).first()
            form = CareerForm(instance=career)
            tiles = career.tiles.all().order_by("order")
            old_order = dict(
                tiles.annotate(
                    str_id=Cast("uid", output_field=TextField())
                ).values_list("str_id", "order")
            )
        context = {
            "tiles": tiles,
            "form": form,
            "career": career,
            "old_order": json.dumps(old_order),
            "career_uid": career_uid,
            "video_id": career.video_id if career and career.video else None,
            "video_name": career.video.name if career and career.video else None,
            "video_url": career.video.cdn_url if career and career.video else None,
        }

        if request.is_ajax():
            html = render_to_string("college_tiles_list.html", context)
            return HttpResponse(html)
        return render(request, self.template_name, context)

    def post(self, request, career_uid):
        """Create a new College."""
        data = request.POST.copy()
        career = Career.objects.filter(
            university=self.university, uid=career_uid
        ).first()

        if career_uid:
            career = Career.objects.filter(
                university=self.university, uid=career_uid
            ).first()
            if data.get("type", None) == "tile_update":
                tile = Tile()
                tile_uid = data.get("tile_uid", None)
                if (
                    tile_uid
                    and Tile.objects.filter(
                        university=self.university, uid=tile_uid
                    ).first()
                ):
                    tile = Tile.objects.filter(
                        university=self.university, uid=tile_uid
                    ).first()
                    order = tile.order
                else:
                    order = (
                        career.tiles.all().all().order_by("-order").first().order + 1
                        if career.tiles.all().exists()
                        else 1
                    )
                tile.order = order
                active = data.get("tile_active")
                tile.active = True if active == "true" else False
                tile_form = TileForm(data, request.FILES, instance=tile)
                if tile_form.is_valid():
                    tile.university = self.university
                    tile = tile_form.save()
                    if tile.url_type == "external":
                        tile.page = None
                    if not data.get("image_url", None) and not request.FILES.get(
                        "image"
                    ):
                        tile.image = None
                    if not tile.active:
                        tile.modified_by_page = "deactive"
                    else:
                        tile.modified_by_page = None
                    tile.save()
                    career.tiles.add(tile)
                else:
                    return HttpResponse(json.dumps('{result: "Error"}'))
                return HttpResponse(json.dumps('{result: "Success"}'))
            elif data.get("type", None) == "order_update":
                order_data = json.loads(data.get("values"))
                for tile in Tile.objects.filter(uid__in=order_data.keys()):
                    tile.order = order_data[str(tile.uid)]
                    tile.university = self.university
                    tile.save()
        career.university = self.university
        data["university"] = str(self.university.uid)
        form = CareerForm(data, request.FILES, instance=career)
        if form.is_valid():
            career = form.save()
            career.active = True
            career.save()
        else:
            return JsonResponse(form.errors.as_json(), safe=False)
        if (
            data.get("file_id", None)
            and File.objects.filter(
                university=self.university, uid=data.get("file_id")
            ).exists()
        ):
            video = File.objects.get(
                university=self.university, uid=data.get("file_id")
            )
            if data.get("video_title"):
                video.name = data.get("video_title")
            if request.FILES.get("thumbnail_image"):
                video.thumbnail = request.FILES.get("thumbnail_image")
            video.save()
            career.video = video
            career.save()

        tiles = career.tiles.all()
        context = {
            "tiles": tiles,
            "form": form,
            "career": career,
            "career_uid": career_uid,
            "video_id": career.video_id if career.video else None,
            "video_name": career.video.name if career.video else None,
            "video_url": career.video.cdn_url if career.video else None,
        }
        return render(request, self.template_name, context)
